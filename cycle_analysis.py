import os
import pandas as pd
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings("ignore")

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

try:
    import tkinter as tk
    from tkinter import filedialog
    TK_AVAILABLE = True
except ImportError:
    TK_AVAILABLE = False

# ===================== 配置区 =====================
COL_DATE = "送货日期"
COL_GOODS = "商品名称"
COL_UNIT = "订货单位"
COL_PRICE = "售价"
COL_AMOUNT = "订货数量"
SKIP_ROWS = 1
# 输出文件
OUTPUT_FILE = "商品周期对比分析.xlsx"
# ==================================================

def select_files_interactive():
    print("=" * 60)
    print("  商品销量&价格周期性分析工具")
    print("=" * 60)
    if TK_AVAILABLE:
        print("\n📂 正在打开文件选择窗口...")
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        file_paths = filedialog.askopenfilenames(
            title="选择表格文件（可多选）",
            filetypes=[("表格", "*.xlsx *.xls *.csv"), ("Excel", "*.xlsx"), ("CSV", "*.csv")]
        )
        root.destroy()
        if file_paths:
            print(f"✅ 选中 {len(file_paths)} 个文件")
            return list(file_paths)
    print("⚠️ 无图形界面，手动输入路径（多文件逗号分隔/文件夹路径）：")
    input_path = input("路径: ").strip()
    paths = [p.strip() for p in input_path.split(",") if p.strip()]
    all_files = []
    for p in paths:
        if os.path.isdir(p):
            for root, _, files in os.walk(p):
                for f in files:
                    if f.lower().endswith((".xlsx", ".xls", ".csv")):
                        all_files.append(os.path.join(root, f))
            print(f"📂 文件夹读取到 {len(all_files)} 个表格")
        elif os.path.isfile(p):
            all_files.append(p)
        else:
            print(f"⚠️ 无效路径跳过：{p}")
    if not all_files:
        raise FileNotFoundError("未找到有效表格文件")
    return all_files


def input_cycle_config():
    print("\n" + "=" * 60)
    print("  周期设置（统一起始日期 + 固定天数周期）")
    print("=" * 60)

    while True:
        start_input = input("\n请输入周期起始日期（格式：2026-01-01，留空用数据最早日期）: ").strip()
        if not start_input:
            start_date = None
            print("✅ 将使用数据中最早的日期作为起始")
            break
        try:
            start_date = pd.to_datetime(start_input)
            print(f"✅ 起始日期：{start_date.strftime('%Y年%m月%d日')}")
            break
        except:
            print("❌ 日期格式不对，请用 YYYY-MM-DD 格式，比如 2026-01-01")

    while True:
        days_input = input("\n请输入每个周期的天数（如7=周、15=半月、30=月）: ").strip()
        try:
            cycle_days = int(days_input)
            if cycle_days > 0:
                print(f"✅ 周期长度：{cycle_days} 天")
                break
            print("❌ 天数必须大于0")
        except ValueError:
            print("❌ 请输入有效的数字")

    return start_date, cycle_days


def read_tables(file_paths):
    all_df = []
    for fp in file_paths:
        if fp.lower().endswith(".csv"):
            df = pd.read_csv(fp, skiprows=SKIP_ROWS)
        else:
            df = pd.read_excel(fp, engine="openpyxl", skiprows=SKIP_ROWS)
        all_df.append(df)
        print(f"读取：{os.path.basename(fp)} 共{len(df)}行")
    total = pd.concat(all_df, ignore_index=True)
    print(f"\n合并总数据：{len(total)}行")
    print(f"列名：{list(total.columns)}")
    return total


def create_uniform_cycles(df, user_start_date, cycle_days):
    df[COL_DATE] = pd.to_datetime(df[COL_DATE], errors="coerce")
    df = df.dropna(subset=[COL_DATE])

    if df.empty:
        raise ValueError("时间列无有效数据，请检查表名/跳过行数")

    data_min = df[COL_DATE].min()
    data_max = df[COL_DATE].max()
    print(f"\n📅 数据时间范围：{data_min.date()}  ~  {data_max.date()}")

    if user_start_date is None:
        start_date = data_min.normalize()
    else:
        start_date = user_start_date.normalize()
        if start_date > data_max:
            raise ValueError(f"起始日期 {start_date.date()} 晚于数据最晚日期 {data_max.date()}")

    total_days = (data_max - start_date).days
    full_cycle_count = total_days // cycle_days

    if full_cycle_count == 0:
        raise ValueError(f"数据跨度不足 {cycle_days} 天，无法形成一个完整周期")

    end_date = start_date + timedelta(days=full_cycle_count * cycle_days)

    print(f"\n📊 周期划分方案：")
    print(f"   起始日期：{start_date.strftime('%Y年%m月%d日')}")
    print(f"   周期长度：{cycle_days} 天")
    print(f"   完整周期数：{full_cycle_count} 个")
    print(f"   截止日期：{end_date.strftime('%Y年%m月%d日')}（不含当日）")

    df = df[(df[COL_DATE] >= start_date) & (df[COL_DATE] < end_date)]
    print(f"   过滤后有效数据：{len(df)} 行")

    def get_cycle_label(dt):
        offset_days = (dt - start_date).days
        cycle_idx = offset_days // cycle_days
        cycle_start = start_date + timedelta(days=cycle_idx * cycle_days)
        cycle_end = cycle_start + timedelta(days=cycle_days - 1)
        return f"{cycle_start.month:02d}月{cycle_start.day:02d}日~{cycle_end.month:02d}月{cycle_end.day:02d}日"

    df["周期"] = df[COL_DATE].apply(get_cycle_label)

    cycle_list = sorted(df["周期"].unique())
    print(f"\n📋 全部 {len(cycle_list)} 个周期：")
    for i, c in enumerate(cycle_list, 1):
        print(f"   周期{i}：{c}")

    return df, cycle_list


def analysis_detail(df):
    """按商品+周期聚合"""
    group = df.groupby([COL_GOODS, "周期"]).agg(
        订货单位=(COL_UNIT, "first"),
        售价=(COL_PRICE, "first"),
        订货数量=(COL_AMOUNT, "sum"),
        记录条数=(COL_DATE, "count")
    ).reset_index()

    group["订货数量"] = group["订货数量"].round(2)
    group["售价"] = group["售价"].round(2)

    group = group.sort_values([COL_GOODS, "周期"])

    # 环比计算
    group["销量环比"] = group.groupby(COL_GOODS)["订货数量"].diff().round(2)
    group["价格环比"] = group.groupby(COL_GOODS)["售价"].diff().round(2)
    group["销量波动"] = (group["销量环比"] / group.groupby(COL_GOODS)["订货数量"].shift(1) * 100).round(2)
    group["价格波动"] = (group["价格环比"] / group.groupby(COL_GOODS)["售价"].shift(1) * 100).round(2)

    return group


def export_beautiful_excel(df, output_path):
    """美化输出Excel：列宽、表头、边框、对齐、百分比格式"""
    wb = Workbook()
    ws = wb.active
    ws.title = "周期对比分析"

    # 定义样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="微软雅黑", size=10)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 写入表头
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, row_data in enumerate(df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.border = thin_border

            # 百分比列（销量波动、价格波动）
            col_name = headers[col_idx - 1]
            if "波动" in col_name:
                if pd.notna(value) and value != "":
                    cell.value = value / 100
                    cell.number_format = '0.00%'
                cell.alignment = cell_alignment
            elif col_name in ["商品名称", "周期", "订货单位"]:
                cell.alignment = left_alignment
            else:
                cell.alignment = cell_alignment

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_value in row:
                if cell_value:
                    length = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in str(cell_value))
                    if length > max_length:
                        max_length = length
        ws.column_dimensions[col_letter].width = max(max_length + 4, 10)

    # 行高
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = "A2"

    # 隔行底色
    light_blue_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for row_idx in range(2, len(df) + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.fill.start_color.rgb == "00000000":
                    cell.fill = light_blue_fill

    wb.save(output_path)
    print(f"✅ 美化表格已保存：{output_path}")


if __name__ == "__main__":
    try:
        file_list = select_files_interactive()
        start_date, cycle_days = input_cycle_config()

        print("\n" + "=" * 60)
        print("  正在读取数据...")
        print("=" * 60)
        raw_data = read_tables(file_list)

        cycle_data, cycle_order = create_uniform_cycles(raw_data, start_date, cycle_days)
        detail_result = analysis_detail(cycle_data)

        output_path = os.path.abspath(OUTPUT_FILE)
        export_beautiful_excel(detail_result, output_path)

        print("\n" + "=" * 60)
        print("✅ 分析完成！")
        print("=" * 60)

        goods_count = detail_result[COL_GOODS].nunique()
        print(f"\n📊 分析概览：")
        print(f"   商品数量：{goods_count} 个")
        print(f"   周期数量：{len(cycle_order)} 个")
        print(f"   输出文件：{output_path}")

        print(f"\n📋 表格字段：")
        print(f"   商品名称、周期、订货单位、售价、订货数量")
        print(f"   销量环比、价格环比、销量波动%、价格波动%")

        input("\n按回车键退出程序")

    except Exception as err:
        print(f"\n❌ 运行报错：{err}")
        import traceback
        traceback.print_exc()
        print("\n💡 排查建议：")
        print("   1. 检查表头名称是否正确")
        print("   2. 确认送货日期格式正确")
        print("   3. 确保售价和订货数量是数字")
        input("按回车键关闭")